import os
import openpyxl as pyxl
from datetime import date

def updateSchedule(roster, locos, settings):
	wb = pyxl.load_workbook(settings.schedulebase)
	ws = wb.active

	changed = False
	for i in range(4,50):

		tx = "F%d" % i
		trainid = ws[tx].value
		if trainid is None:
			continue

		lx = "G%d" % i
		locoid = ws[lx].value
		if str(locoid).startswith("=+"):
			locoid = ws[locoid[2:]].value

		dx = "H%d" % i

		if locoid is None:
			tInfo = roster.getTrain(str(trainid))
			lid = tInfo["loco"]
			if lid is None:
				locoid = None
				desc = None
			else:
				desc = locos.getLoco(lid)
				locoid = int(lid)
				if desc is None:
					desc = None

			if locoid is not None:
				ws[lx] = locoid
				ws[dx] = desc
				changed = True

	if changed:
		try:
			os.mkdir(settings.scheduledir)
		except FileExistsError:
			pass

		td = date.today()
		base, ext = os.path.splitext(settings.schedulebase)
		bn = "%s%02d-%02d-%02d%s" % (base, td.year%100, td.month, td.day, ext)
		ofn = os.path.join(settings.scheduledir, bn)
		
		wb.save(ofn)
		return bn
		
	else:
		return None

